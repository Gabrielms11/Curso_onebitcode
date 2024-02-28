from openpyxl import load_workbook

# 1-Lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatório"]

# 2-Referências das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
print(min_column)
print(max_column)